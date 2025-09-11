!pip install telebot
!pip install openpyxl

import telebot
from telebot import types
import csv
import os

token = "8296673968:AAEC9Y7tiddFDwQdlPFnEz76SdXORIGVZfI"
bot = telebot.TeleBot(token)

from openpyxl import Workbook, load_workbook
import os
import datetime

@bot.message_handler(commands=["start"])
def start_msg(message):
    store_interaction_data(message)  # Store initial user data
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False) # Changed to False
    btn1 = types.KeyboardButton("Consultation & personalized help")
    btn2 = types.KeyboardButton("Job openings/referrals")
    btn3 = types.KeyboardButton("Get free PDF")
    btn4 = types.KeyboardButton("End chat")
    btn5 = types.KeyboardButton("Contact Us")
    markup.add(btn1, btn2, btn3, btn4, btn5)
    bot.send_message(message.chat.id,
        "Hey user, Gerry's Bot this side 👋\n\nWelcome to our Data Career Support bot.\n\nPlease choose one of the following:",
        reply_markup=markup
    )

@bot.message_handler(func=lambda message: message.text == "Consultation & personalized help")
def handle_consultation(message):
    store_interaction_data(message, "Clicked Button", message.text) # Store the clicked button text
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False) # Changed to False
    options = [
        "🔹 Not getting interviews",
        "🔹 Not getting shortlisted",
        "🔹 Low salary / stuck role",
        "🔹 Confused about upskilling",
        "🔹 Other"
    ]
    for option in options:
        markup.add(option)
    bot.send_message(
        message.chat.id,
        "Before we begin, could you share your biggest challenge right now?\n(Select one)",
        reply_markup=markup
    )

@bot.message_handler(func=lambda message: message.text in [
    "🔹 Not getting interviews",
    "🔹 Not getting shortlisted",
    "🔹 Low salary / stuck role",
    "🔹 Confused about upskilling",
    "🔹 Other"
])
def handle_challenge_response(message):
    user_response = message.text  # Capture the user's response
    store_interaction_data(message, "Challenge Response", user_response) # Store the response in the file

    # Use InlineKeyboardMarkup for the Topmate link
    markup_topmate = types.InlineKeyboardMarkup()
    btn_topmate = types.InlineKeyboardButton("Book Your 1:1 Consult Call", url="https://topmate.io/gerryson/870539")
    markup_topmate.add(btn_topmate)

    bot.send_message(
        message.chat.id,
        f"""Thanks for sharing! 🙌

Here’s how we can support you 🚀

Gerryson Mehta has 7+ years of experience in data analytics across companies like Coinbase, Mobikwik, and Tech Mahindra.
He specializes in SQL, Tableau, Power BI, and Snowflake—helping professionals transition into higher-paying analytics roles and secure global opportunities.

✨ Use code FIRST1000 to get 90% off your first call! ✨""",
        reply_markup=markup_topmate # Add the inline keyboard for Topmate link
    )

    followup_markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False) # Changed to False
    btn1 = types.KeyboardButton("Consultation & personalized help")
    btn2 = types.KeyboardButton("Job openings/referrals")
    btn3 = types.KeyboardButton("Get free PDF")
    btn4 = types.KeyboardButton("End chat")
    btn5 = types.KeyboardButton("Contact Us")
    followup_markup.add(btn1, btn2, btn3, btn4, btn5)
    bot.send_message(
        message.chat.id,
        "Do you have any other queries you'd like help with?\nFeel free to explore more or end the chat below 👇",
        reply_markup=followup_markup
    )
    bot.send_message(
        message.chat.id,
        "Thanks for connecting! 🙏\nYou can explore more resources at:\n🌐 www.gerrysonmehta.com"
    )

@bot.message_handler(func=lambda message: message.text == "Job openings/referrals")
def handle_jobs(message):
    store_interaction_data(message, "Clicked Button", message.text) # Store the clicked button text
    markup = types.InlineKeyboardMarkup()
    btn = types.InlineKeyboardButton("Join WhatsApp Group", url="https://whatsapp.com/channel/0029VamouNm5Ejy6enHyEd29")
    markup.add(btn)
    bot.send_message(
        message.chat.id,
        "Great! 🎯 Tap below to join our WhatsApp community for curated job openings and referrals.",
        reply_markup=markup
    )

@bot.message_handler(func=lambda message: message.text == "Get free PDF")
def send_pdf_link(message):
    store_interaction_data(message, "Clicked Button", message.text) # Store the clicked button text
    markup = types.InlineKeyboardMarkup()
    btn = types.InlineKeyboardButton("📘 Download Free PDF", url="https://docs.google.com/document/d/e/2PACX-1vTOhSl0g3Q1K_44w5OJFlyBDkOEraufV3sxtojvuQZeIE7S_ptwk0FGjfMi2mohSJ5qgt3-Tw3KbH48/pub")
    markup.add(btn)
    bot.send_message(
        message.chat.id,
        "Here’s your free resource to help you level up in data analytics! 🚀\nTap below to download:",
        reply_markup=markup
    )

@bot.message_handler(func=lambda message: message.text == "End chat")
def handle_end_chat(message):
    store_interaction_data(message, "Clicked Button", message.text) # Store the clicked button text
    bot.send_message(
        message.chat.id,
        "Chat ended ✅\nFeel free to restart anytime by typing /start.\nWishing you success ahead! 🚀"
    )

@bot.message_handler(func=lambda message: message.text == "Contact Us")
def handle_contact_us(message):
    store_interaction_data(message, "Clicked Button", message.text) # Store the clicked button text
    markup = types.InlineKeyboardMarkup()
    btn = types.InlineKeyboardButton("📬 Contact Us Form", url="https://forms.gle/E3hs5TrJuT7zVGMZ6")
    markup.add(btn)
    bot.send_message(
        message.chat.id,
        "Tap below to reach out to us:",
        reply_markup=markup
    )





!pip install pytz

import pytz
import datetime
from openpyxl import Workbook, load_workbook
import os
from telebot import types # Make sure types is imported

def store_interaction_data(message, column_name=None, data=None):
    print("store_interaction_data function called") # Debug print
    user_id = message.chat.id
    # Get full name by combining first and last names
    first_name = message.from_user.first_name or ""
    last_name = message.from_user.last_name or ""
    name = f"{first_name} {last_name}".strip() or "N/A"
    username = message.from_user.username or "N/A"
    # Get current time in UTC and convert to IST
    utc_now = datetime.datetime.now(datetime.timezone.utc)
    ist_timezone = pytz.timezone('Asia/Kolkata')
    ist_now = utc_now.astimezone(ist_timezone)
    timestamp = ist_now.strftime("%Y-%m-%d %H:%M:%S") # Added time for better tracking

    mobile = "N/A"
    email = "N/A"
    file_path = "/content/drive/My Drive/Colab Notebooks/telegram_user_data_all.xlsx" # Updated filename
    print(f"File path: {file_path}") # Debug print

    drive_dir = "/content/drive/My Drive/Colab Notebooks/"
    if not os.path.exists(drive_dir):
        os.makedirs(drive_dir)
        print(f"Created directory: {drive_dir}") # Debug print

    header = [
    "User ID", "Name", "Username", "Timestamp", "Mobile", "Email",
    "Challenge Response", "Clicked Button", "Gender", "Location", "Language", "Referral Source"
]

    if not os.path.exists(file_path):
        print(f"File not found, creating new workbook: {file_path}") # Debug print
        wb = Workbook()
        ws = wb.active
        ws.title = "User Info"
        ws.append(header)
    else:
        print(f"File found, loading workbook: {file_path}") # Debug print
        try:
            wb = load_workbook(file_path)
            ws = wb["User Info"]
            existing_header = [cell.value for cell in ws[1]]
            for col_name in header:
                if col_name not in existing_header:
                    ws.cell(row=1, column=len(existing_header) + 1, value=col_name)
                    existing_header.append(col_name)
                    print(f"Added missing column: {col_name}") # Debug print
        except Exception as e:
            print(f"Error loading workbook: {e}") # Debug print
            # If loading fails, create a new workbook
            print("Creating a new workbook due to loading error.")
            wb = Workbook()
            ws = wb.active
            ws.title = "User Info"
            ws.append(header)


    # Create a new row with initial data
    new_row_data = [user_id, name, username, timestamp, mobile, email, "N/A", "N/A", "N/A", "N/A", "N/A", "N/A"]

    # Update the specific column if provided
    if column_name and data is not None:
        try:
            col_index = header.index(column_name)
            new_row_data[col_index] = data
        except ValueError:
            print(f"Warning: Column '{column_name}' not found in header.")


    ws.append(new_row_data)
    print(f"Appended data for user ID: {user_id}") # Debug print

    try:
        wb.save(file_path)
        print("Workbook saved successfully") # Debug print
    except Exception as e:
        print(f"Error saving workbook: {e}") # Debug print


@bot.message_handler(commands=["start"])
def start_msg(message):
    print("start_msg function called") # Debug print
    store_interaction_data(message)  # Store initial user data on start
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
    btn1 = types.KeyboardButton("Consultation & personalized help")
    btn2 = types.KeyboardButton("Job openings/referrals")
    btn3 = types.KeyboardButton("Get free PDF")
    btn4 = types.KeyboardButton("End chat")
    btn5 = types.KeyboardButton("Contact Us")
    markup.add(btn1, btn2, btn3, btn4, btn5)
    bot.send_message(message.chat.id,
        "Hey user, Gerry's Bot this side 👋\n\nWelcome to our Data Career Support bot.\n\nPlease choose one of the following:",
        reply_markup=markup
    )
    print("start message sent") # Debug print

bot.polling()

